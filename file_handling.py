from json import load
import openpyxl

def get_school_to_link(file_name:str, county:str):
    """
    Extracts dictionary that maps school to link from json indicated by file_name
    @param file_name
    """
    file = open("./hs_links/"+file_name)
    school_to_link=load(file)[county]
    file.close()
    return school_to_link

def write_to_excel_file(contact_info:dict,county:str,file_name:str):
    """
    Writes the content of contact_info to an Excel file 
    @param contact_info : dictionary mapping school name indicated by school to a dictionary mapping school staff member name to email
    @param county
    @param file_name
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = county
    sheet.cell(row=1, column=1).value='School'
    sheet.cell(row=1,column=2).value='Counselor Name'
    sheet.cell(row=1, column=3).value='Email'
    i = 2
    for school in contact_info:
        for couns in contact_info[school]:
            sheet.cell(row=i,column=1).value=school
            sheet.cell(row=i,column=2).value=couns
            sheet.cell(row=i,column=3).value=contact_info[school][couns]
            i+=1
    wb.save(file_name)