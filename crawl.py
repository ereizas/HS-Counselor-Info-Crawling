import requests
from bs4 import BeautifulSoup
import re
import openpyxl
from time import sleep,time
from random import randint
from json import dump, load
from search import google_search

def get_state_hs_links(state_hs_lst_url:str,county_to_retrieve:str=None):
    """
    Parses wikipedia page to create a dictionary that maps county to a dictionary that maps school name to school url and writes to a json file with 
    the created dictionary hs_links
    @param state_hs_lst_url : wikipedia url for list of high schools in a certain state
    @param county_to_retrieve : str of a county in PA
    """
    req = requests.get(state_hs_lst_url)
    soup = BeautifulSoup(req.text,'html.parser')
    hs_links = dict()
    schools_html = soup.find_all(re.compile('span|div'),attrs={'class':re.compile('mw-headline|div-col')})
    len_schools_html = len(schools_html)
    i=0
    if county_to_retrieve:
        while i<len_schools_html and (schools_html[i].name!='span' or county_to_retrieve not in schools_html[i].text):
            i+=1
        i+=1
    while i<len_schools_html:
        county = None if not county_to_retrieve else county_to_retrieve
        if not county_to_retrieve:
            county = schools_html[i].text
            if 'County' not in county:
                break
            i+=1
        hs_links[county]=dict()
        #Added to query to ensure the correct school is retrieved
        city = None
        while i<len_schools_html and (schools_html[i].name=='div' or (schools_html[i].name=='span' and schools_html[i].parent.name!='h2')):
            if schools_html[i].name=='span' and schools_html[i].find('a'):
                city = schools_html[i].text
            elif schools_html[i].name=='div':
                schools=schools_html[i].find_all('li')
                start_time = None
                for school in schools:
                    external_link=school.find('a',attrs={'class':'external text'})
                    if external_link:
                        school_text=school.text
                        comma_ind = school.text.find(',')
                        if comma_ind==-1:
                            comma_ind = len(school.text)
                        hs_links[county][school_text[:comma_ind]]=external_link.get('href')
                    else:
                        query = school.text
                        if ',' not in query:
                            if city:
                                query+=' ' + city
                            else:
                                query+=' ' +county
                        if start_time!=None:
                            sleep_time=randint(30, 40)-(time()-start_time)
                            print(f"Sleeping for {sleep_time} seconds")
                            sleep(sleep_time)
                        print(query)
                        start_time=time()
                        goog_srch_res = google_search(query)
                        if goog_srch_res:
                            comma_ind = school.text.find(',')
                            if comma_ind==-1:
                                comma_ind = len(school.text)
                            hs_links[county][school.text[:comma_ind]]=goog_srch_res
            i+=1
        if county_to_retrieve:
            break
    print(hs_links)
    with open('hs_links.json' if not county_to_retrieve else county_to_retrieve.lower()+'_hs_links'+'.json','w') as file:
        dump(hs_links,file)
    file.close()

def get_contacts_from_sprdsheet(soup:BeautifulSoup,job_col:str,name_cols:list[str],email_col:str,contact_info:dict,school:str,name_rvrs_order=False):
    """
    Parses contacts of high school Special Ed staff and Counselors into a dictionary from a certain format of spreadsheet within HTML of the school website indicated by school
    @param soup
    @param job_col : str column num where the job of the staff member is located
    @param name_cols : list of columns where the full name or first and last name of the staff member is located
    @param email_col : str column where the email of the staff member is located
    @param contact_info : dictionary mapping school name indicated by school to a dictionary mapping school staff member name to email
    @param school : name of the school
    @param name_rvrs_order : boolean indicating whether the names are in reverse order in a single column
    """
    if school=='Shoemaker':
        soup=soup.find('table',attrs={'id':'tablepress-263'})
    rows = soup.find_all('tr',attrs={'class':re.compile("row-([2-9]|[1-9]\d{1,}) ?(even|odd)?")})
    for row in rows:
        td_tag = row.find('td',attrs={'class':'column-'+job_col})
        if td_tag:
            row_text=td_tag.text
            if ('Counsel' in row_text or 'SPED' in row_text or 'Special Ed' in row_text or 'Autis' in row_text) and 'MS' not in row_text and 'Bilingual' not in row_text:
                names = row.find('td',attrs={'class':'column-'+name_cols[0]}).text.split('\n')
                for i in range(1,len(name_cols)):
                    new_names = row.find('td',attrs={'class':'column-'+name_cols[i]}).text.split('\n')
                    for j in range(len(names)):
                        names[j]+=' '+new_names[j]
                emails = row.find('td',attrs={'class':'column-'+email_col}).text.split('\n')
                i = 0
                for i in range(len(names)):
                    if emails[i]:
                        if name_rvrs_order:
                            if '(' not in names[i]:
                                names[i]=names[i][names[i].find(',')+2:]+' '+names[i][:names[i].find(',')]
                            else:
                                names[i]=names[i][:names[i].find('(')]
                        emails[i]=emails[i].replace('[dot]','.')
                        emails[i]=emails[i].replace('[at]','@')
                        contact_info[school][names[i]]=emails[i]

def get_contacts_from_ul_tags(soup:BeautifulSoup,school:str,header_num:str,contact_info:dict,title_included=False):
    """
    Parses ul tags of a certain format to retrieve contact info for counselors of the school indicated by school
    @param soup
    @param school : name of the school
    @param header_num : str number for which type of header the staff member info is located
    @param contact_info : dictionary mapping school name indicated by school to a dictionary mapping school staff member name to email
    @param title_included : whether an honorific is included with the staff member name
    """
    ul_tags=soup.find_all('ul',attrs={'class':None,'id':None})[1:]
    regex = '^[A-Z].[a-z]* [A-Z].[a-z]*'
    if title_included:
        regex='^(Mrs?)|(Ms)|(Dr)\. [A-Z].[a-z]* [A-Z].[a-z]*'
    header_tags=soup.find_all('h'+header_num,string=re.compile(regex))
    for i in range(len(ul_tags)):
        header_txt = header_tags[i].text
        colon_ind = header_txt.find(':')
        comma_ind = header_txt.find(',')
        len_text = len(header_txt)
        if colon_ind==-1:
            colon_ind=len_text
        if comma_ind==-1:
            comma_ind=len_text
        end_ind = colon_ind if colon_ind<comma_ind else comma_ind
        contact_info[school][header_txt[:end_ind]]=ul_tags[i].find('a').text

def get_contacts_from_name_pos_class_div_tags(soup:BeautifulSoup,contact_info:dict,school:str):
    """
    Parses div tags that have the classes name-position and email-phone for contact info of counselors of the school indicated by school
    @param soup
    @param contact_info : dictionary mapping school name indicated by school to a dictionary mapping school staff member name to email
    @param school : name of the school
    """
    div_tags=soup.find_all('div',attrs={'class':re.compile('name-position|email-phone')})
    num_tags = len(div_tags)
    for i in range(0,num_tags,2):
        name_txt = div_tags[i].text
        if 'Counselor' in name_txt:
            contact_info[school][div_tags[i].find('a').text.replace('.\n\t','. ').replace('\n','').replace('\t','')]=div_tags[i+1].find('a').text

def get_contacts_from_p_tags(soup:BeautifulSoup,contact_info:dict,school:str):
    """
    Parses p tags of a certain format to extract contact info of Special Education specialists and counselors of the school indicated by school
    @param soup
    @param contact_info : dictionary mapping school name indicated by school to a dictionary mapping school staff member name to email
    @param school : name of the school
    """
    p_tags = soup.find_all('p')
    for tag in p_tags:
        a_tag=tag.find('a',attrs={'href':re.compile('([A-Za-z])*@([A-Za-z0-9])*.org')})
        tag_txt=tag.text
        if a_tag and ('Counselor' in tag_txt or 'Special Ed' in tag_txt) and 'LS' not in tag_txt:
            contact_info[school][a_tag.text]=a_tag.get('href').strip('mailto:')

def get_contacts_from_staff_info_blocks(school_to_link,school,job_keywords,contact_info):
    i=1
    div_tags=1
    while div_tags:
        soup=get_soup(school_to_link[school],'/staff?page_no='+str(i))
        div_tags=soup.find_all('div',attrs={'class':'staff-info'})
        for tag in div_tags:
            title = tag.find('div',attrs={'class':'title'}).text
            if any([keyword in title for keyword in job_keywords]):
                contact_info[school][tag.find('div',attrs={'class':'name'}).text.strip('\n ')]=tag.find('a').text.strip('\n ')
        i+=1

def is_hs_staff(txt:str):
    """
    Checks if the staff in the given text of a certain format is a high school staff member
    @txt
    """
    grade_ind = txt.find('Grade')
    if grade_ind==-1:
        return True
    i=grade_ind+len('Grade ')+1
    grade = ''
    while not txt[i].isnumeric():
        i+=1
    while i<len(txt) and txt[i].isnumeric():
        grade+=txt[i]
        i+=1
    if grade=='K':
        return False
    grade = int(grade)
    return grade>8

def get_soup(url:str,suff:str=''):
    """
    Makes the request and returns the corresponding BeautifulSoup object
    @param url : base link for the school
    @param suff : suffix which leads to counselor or staff page for the school
    """
    req=requests.get(url+suff)
    return BeautifulSoup(req.text,'html.parser')

def get_school_to_link(file_name:str, county:str):
    """
    Extracts dictionary that maps school to link from json indicated by file_name
    @param file_name
    """
    file = open(file_name)
    school_to_link=load(file)[county]
    file.close()
    return school_to_link

def get_phila_county_contacts():
    """
    Retrieves the contact info for high school counselors and special education specialists in Philadelphia County
    """
    contact_info = dict()
    school_to_link=get_school_to_link('philadelphia_school_links.json','Philadelphia County (City of Philadelphia)')
    for school in school_to_link:
        contact_info[school]=dict()
        if school == 'John Bartram High School':
            soup=get_soup(school_to_link[school],'counselors-corner')
            get_contacts_from_ul_tags(soup,school,'3',contact_info,True)
        elif school=='Thomas A. Edison High School':
            soup=get_soup(school_to_link[school],'counselors-corner')
            rows = soup.find_all('tr',attrs={'class':re.compile("row-[2-9] (even|odd)")})
            for row in rows:
                contact_info[school][row.find('td',attrs={'class':'column-1'}).text]=row.find('td',attrs={'class':'column-4'}).text
        elif school=='Samuel Fels High School':
            soup=get_soup(school_to_link[school],'counselors-and-social-work')
            tags=soup.find_all(re.compile('(^p$)|(^h3$)'))
            for i in range(1,len(tags),2):
                name = tags[i].text
                if not bool(re.match('^(Mrs?)|(Ms)|(Dr)\.. [A-Z].[a-z]* -',name)):
                    break
                name = name[:name.find(' -')]
                contact_info[school][name]=tags[i+1].find('em').text
        elif school=='Benjamin Franklin High School':
            soup=get_soup(school_to_link[school],'counselors')
            strong_tags=soup.find_all('strong')[2:]
            for tag in strong_tags:
                tag_txt=tag.text
                contact_info[school][tag_txt[:tag_txt.find('–')-1]]=tag_txt[tag_txt.find('–')+2:]
        elif school=='Kensington High School':
            soup=get_soup(school_to_link[school],'faculty-staff')
            get_contacts_from_sprdsheet(soup,'4',['1'],'5',contact_info,school,True)
        elif school=='Northeast High School':
            soup=get_soup(school_to_link[school],'counseling')
            #first two rows initialize the style
            span_tags = soup.find_all('span',attrs={'style':re.compile('font-weight: (\d)*')})
            contact_info[school][span_tags[8].text]=span_tags[11].text
            #next rows have a different tag
            td_tags = soup.find_all('td',attrs={'style':re.compile('height: (\d)*px;width: (\d)*px')})
            for i in range(11,len(td_tags),5):
                contact_info[school][td_tags[i].text.strip('\xa0')]=td_tags[i+3].text.strip('\xa0')
        elif school=='Penn Treaty School (6-12)':
            soup=get_soup(school_to_link[school],'who-we-are')
            get_contacts_from_sprdsheet(soup,'1',['2'],'3',contact_info,school)
        elif school=='Constitution High School':
            soup=get_soup(school_to_link[school],'staff')
            get_contacts_from_sprdsheet(soup,'1',['2'],'3',contact_info,school)
        elif school=='Roxborough High School':
            soup=get_soup(school_to_link[school],'support-team')
            h3_tags = soup.find_all('h3')
            num_tags=len(h3_tags)
            for i in range(0,num_tags,2):
                if i+1<num_tags:
                    contact_info[school][h3_tags[i].text]=h3_tags[i+1].text
        if school=='South Philadelphia High School':
            soup=get_soup(school_to_link[school],'counselor')
            p_tags=soup.find_all('p')
            i = 0
            num_p_tags=len(p_tags)
            while i<num_p_tags:
                tag_txt = p_tags[i].text
                if 'Counselor' in tag_txt:
                    name = tag_txt
                    name = name[:name.find('(')-1]
                    i+=1
                    a_tag = p_tags[i].find('a')
                    if not a_tag:
                        i+=1
                        tag_txt=p_tags[i].text
                        contact_info[school][name]=tag_txt[tag_txt.find(':')+2:]
                i+=1
        elif school=='George Washington High School':
            soup=get_soup(school_to_link[school],'support-team')
            p_tags=soup.find_all('p',attrs={'style':'text-align: center'})
            for i in range(1,len(p_tags)):
                tag_txt=p_tags[i].text
                if 'Counsel' in tag_txt:
                    a_tag = p_tags[i].find('a')
                    if  a_tag:
                        contact_info[school][tag_txt[:tag_txt.find('\n')]]=a_tag.text.strip('\xa0')
        elif school=='Bodine International Affairs':
            soup=get_soup(school_to_link[school],'faculty-staff')
            li_tags=soup.find_all('li',attrs={'class':None,'id':None})
            for tag in li_tags:
                tag_txt=tag.text
                if 'Counselor' in tag_txt or 'Special Ed' in tag_txt:
                    contact_info[school][tag_txt[:tag_txt.find(',')]]=tag_txt[tag_txt.rfind(',')+2:]
        elif school=='CAPA':
            soup=get_soup(school_to_link[school],'counselor-page')
            a_tags = soup.find_all('a',attrs={'href':re.compile('([A-Za-z])@philasd.org')})
            for tag in a_tags:
                tag_txt=tag.text
                name = tag_txt[6:]
                dot_ind = name.find('.')
                name = name[0]+name[1:dot_ind+2].lower()+name[dot_ind+2]+name[dot_ind+3:].lower()
                contact_info[school][name]=tag.get('href')
        elif school=='Central High School':
            soup=get_soup(school_to_link[school],'parents-2/counselors/')
            td_tags=soup.find_all('td',attrs={'style':'text-align: center'})
            for i in range(3,len(td_tags),2):
                tag_txt=td_tags[i].text
                #' x' should be a valid cutoff for the name assuming names start with a capital letter (i.e. Xavier)
                contact_info[school][tag_txt[:tag_txt.find(' x')]]=tag_txt[tag_txt.find('/')+2:]
        elif school == 'GAMP':
            soup=get_soup(school_to_link[school],'counselors-corner')
            p_tags=soup.find_all('p')
            for tag in p_tags:
                tag_txt = tag.text
                if '@' in tag_txt:
                    #long dash is used not short dash
                    contact_info[school][tag_txt[4:tag_txt.find('–')-1]]=tag_txt[tag_txt.find('–')+2:]
        elif school in ['Hill-Freedman World Academy High School','The LINC']:
            soup=get_soup(school_to_link[school],'parents/faculty-staff')
            get_contacts_from_sprdsheet(soup,'2','1','3',contact_info,school)
        elif school=='Kensington Health Sciences Academy High School':
            soup=get_soup(school_to_link[school],'counseling')
            tags = soup.find_all(re.compile('th|td'),attrs={'class':re.compile('column-\d|^$')})
            for tag in tags:
                tag_txt = tag.text
                begin_ind = tag_txt.find('\n')
                if begin_ind!=-1:
                    end_ind = tag_txt.find('\n',begin_ind+1)
                    job_str = tag_txt[begin_ind:end_ind]
                    if 'Counselor' in job_str or 'Special Ed' in job_str:
                        contact_info[school][tag.find('u').text.strip(' ')]=tag.find('a').text.strip(' ')
        elif school=='Parkway Northwest High School':
            soup=get_soup(school_to_link[school],'counselor-corner')
            info_btn = soup.find('a',attrs={'class':'vc_general vc_btn3 vc_btn3-size-md vc_btn3-shape-rounded vc_btn3-style-modern vc_btn3-color-sandy-brown'})
            contact_info[school][info_btn.text[:info_btn.text.find('-')]]=info_btn.get('href').strip('mailto:')
        elif school in 'Science Leadership Academy':
            soup=get_soup(school_to_link[school],'faculty')
            get_contacts_from_p_tags(soup,contact_info,school)
        elif school in ['Science Leadership Academy at Beeber (6-12)','The Crefeld School']:
            soup=get_soup(school_to_link[school],'faculty-and-staff')
            get_contacts_from_p_tags(soup,contact_info,school)
        elif school=='Murrell Dobbins Vocational School':
            soup=get_soup(school_to_link[school],'about-us/staff-list')
            p_tags=soup.find_all('p')
            i = 0
            num_p_tags=len(p_tags)
            while i<num_p_tags:
                tag_txt = p_tags[i].text
                name = tag_txt
                i+=1
                a_tag = p_tags[i].find('a')
                if not a_tag:
                    i+=1
                    tag_txt=p_tags[i].text
                    if 'Counsel' in p_tags[i-1].text or 'Special Ed' in p_tags[i-1].text or 'SPED' in p_tags[i-1].text:
                        contact_info[school][name]=p_tags[i].find('a').text
                i+=1
        elif school=='Philadelphia Military Academy':
            soup=get_soup(school_to_link[school],'staff')
            get_contacts_from_sprdsheet(soup,'2','1','3',contact_info,school)
        elif school=='Kensington Creative & Performing Arts High School':
            soup=get_soup(school_to_link[school],'staff')
            a_tags=soup.find_all('a',attrs={'data-vc-container':'.vc_tta','data-vc-tabs':''})
            panels = soup.find_all('div',attrs={'class':'vc_tta-panel-body'})
            panels = [panels[i] for i in range(len(a_tags)) if 'Support' in a_tags[i].find('span').text]
            for panel in panels:
                h2_tags=panel.find_all('h2')
                h3_tags=panel.find_all('h3')
                div_tags=panel.find_all('div',attrs={'class':'vc_toggle_content'})
                for i in range(len(h2_tags)):
                    tag_txt = h2_tags[i].text
                    if 'Counselor' in tag_txt or 'SPED' in tag_txt or 'Special Ed' in tag_txt:
                        contact_info[school][h3_tags[i].text]=div_tags[i].find('p').text
        elif school=='Jules E. Mastbaum Technical High School':
            soup=get_soup(school_to_link[school],'contact-us')
            get_contacts_from_sprdsheet(soup,'2','3','4',contact_info,school)
        elif school=='High School of the Future':
            soup=get_soup(school_to_link[school],'staff')
            get_contacts_from_sprdsheet(soup,'3',['1','2'],'4',contact_info,school)
        elif school=='Randolph Technical High School':
            soup=get_soup(school_to_link[school],'about-us/staff')
            li_tags=soup.find_all('li',attrs={'class':None,'id':None})
            for i in range(len(li_tags)):
                strong_tag = li_tags[i].find('strong')
                if strong_tag and 'Counselor' in strong_tag.text:
                    j=i+1
                    strong_tag = li_tags[j].find('strong')
                    while not strong_tag:
                        tag_txt=li_tags[j].text
                        separator_ind = tag_txt.find('(')
                        contact_info[school][tag_txt[:separator_ind-1]]=tag_txt[separator_ind+1:-1]
                        j+=1
                        strong_tag = li_tags[j].find('strong')
                    break
        if school=='Swenson Arts and Technology High School':
            soup=get_soup(school_to_link[school],'counselors-corner')
            get_contacts_from_ul_tags(soup,school,'4',contact_info)
        elif school=='New Foundations Charter School':
            soup=get_soup(school_to_link[school],'counseling')
            div_tags=soup.find_all('div',attrs={'class':'col-12 col-md-6 vw-team-item-wrap text-center'})
            for tag in div_tags:
                tag_txt=tag.text
                if 'Counselor' in tag_txt and 'Grade' not in tag_txt:
                    contact_info[school][tag.find('h5').text]=tag.find('a').get('href').strip('mailto:')
        elif school=='Philadelphia Performing Arts Charter School':
            soup=get_soup(school_to_link[school],'our-families/support-services')
            p_tags = soup.find_all('p')
            for tag in p_tags:
                tag_txt=tag.text
                #manual function to check for grade
                if is_hs_staff(tag_txt) and ('Counselor' in tag_txt or 'Special Ed' in tag_txt):
                    strong_txt = tag.find('strong').text
                    lines = tag_txt.split('\n')
                    contact_info[school][strong_txt[:strong_txt.find('\n')].replace('\xa0',' ')]=lines[-1]
        elif school=='Springside Chestnut Hill Academy':
            soup=get_soup(school_to_link[school],'support')
            h3_tags=soup.find_all('h3')
            a_tags=soup.find_all('a',string=re.compile('([A-Za-z])*@([A-Za-z0-9])*.org'))
            i = 0
            while i < len(h3_tags):
                h3_txt = h3_tags[i].text
                if 'Upper School' in h3_txt:
                    contact_info[school][h3_txt[:h3_txt.find(',')]]=a_tags[i].text
                i+=1
        elif school=='Archbishop Ryan High School':
            soup=get_soup(school_to_link[school],'apps/staff/')
            get_contacts_from_name_pos_class_div_tags(soup,contact_info,school)
        elif school=='Father Judge High School':
            soup=get_soup(school_to_link[school],'apps/staff/')
            get_contacts_from_name_pos_class_div_tags(soup,contact_info,school)
        elif school=='St. Hubert Catholic High School for Girls':
            soup=get_soup(school_to_link[school],'apps/staff/')
            get_contacts_from_name_pos_class_div_tags(soup,contact_info,school)
        elif school=='Mastery Charter Schools (Gratz, Lenfest, Pickett, Shoemaker, Thomas, Hardy Williams)':
            #Mastery Charter has several schools to account for programmatically with the same structure
            req=requests.get(school_to_link[school])
            soup = BeautifulSoup(req.text,'html.parser')
            a_tags=soup.find_all('a',string=re.compile('-12$'))
            for tag in a_tags:
                campus_name = ''
                for char in tag.text:
                    if not char.isnumeric():
                        campus_name+=char
                    else:
                        campus_name=campus_name.strip(' ')
                        break
                campus_name+=' Mastery Charter School Campus'
                contact_info[campus_name]=dict()
                temp_req = requests.get(tag.get('href'))
                soup = BeautifulSoup(temp_req.text,'html.parser')
                get_contacts_from_sprdsheet(soup,'3',['1','2'],'4',contact_info,campus_name)
        elif school=='Mariana Bracetti Academy Charter School':
            soup=get_soup(school_to_link[school],'our-team')
            div_tags = soup.find_all('div',attrs={'class':re.compile('__column$')})[2:]
            for tag in div_tags:
                tag_txt=tag.text
                if 'Special Ed' in tag_txt:
                    a_tag = tag.find('a')
                    contact_info[school][a_tag.text]=a_tag.get('href').strip('mailto:')
        elif school=='Freire Charter School':
            soup=get_soup(school_to_link[school],'staff-directory/')
            div_tags=soup.find_all('div',attrs={'class':'wp-block-genesis-blocks-gb-column gb-block-layout-column gb-is-vertically-aligned-center'})
            for tag in div_tags:
                job_txt = tag.find('p',attrs={'class':None}).text
                if 'Support Service' in job_txt or 'Psychologist' in job_txt:
                    contact_info[school][tag.find('p').text]=tag.find('a').text
    return contact_info

def get_pittsburgh_contacts():
    """
    Retrieves contact info for Pittsburgh high school counselors and special education specialists
    """
    contact_info=dict()
    school_to_link=get_school_to_link('pittsburgh_hs_links.json','Pittsburgh')
    for school in school_to_link:
        contact_info[school]=dict()
        if school=='Brashear High School':
            soup=get_soup(school_to_link[school],'staff')
            tr_tags=soup.find_all('tr')
            for tag in tr_tags:
                staff_info=tag.find_all('td')
                job_txt=staff_info[1].text
                if 'Counselor' in job_txt and 'Transition' not in job_txt:
                    contact_info[school][staff_info[0].text]=staff_info[2].text
        elif school=='City Charter High School':
            soup=get_soup(school_to_link[school],'meet-our-staff#support')
            div_tags = soup.find_all('div','col-md-5')
            for tag in div_tags:
                tag_txt = tag.text
                if 'Counselor' in tag_txt or 'Special Ed' in tag_txt:
                    strong_tag = tag.find('strong')
                    b_tag = tag.find('b')
                    if b_tag:
                        contact_info[school][b_tag.text[:b_tag.text.find(',')]]=tag.find('a').text
                    elif strong_tag:
                        contact_info[school][strong_tag.text[:strong_tag.text.find(',')]]=tag.find('a').text
        elif school=='Pittsburgh Obama 6-12':
            soup=get_soup('https://www.pghschools.org/domain/823')
            strong_tags=soup.find_all('strong')
            for tag in strong_tags:
                tag_txt=tag.text
                #space at the end of counselor is to prevent the heading "Counselors and Social Workers" from being parsed
                #Nurse is ignored bc it is joined w a relevant contact
                if ('Counselor ' in tag_txt or 'Learning Support' in tag_txt or 'Psychologist' in tag_txt) and ' Nurse' not in tag_txt:
                    dash_ind = tag_txt.find('-')
                    contact_info[school][tag_txt[:dash_ind-1].replace('\xa0',' ')]=tag_txt[tag_txt.find('-', dash_ind+1)+2:]
        elif school=='The University School':
            soup=get_soup(school_to_link[school],'faculty/')
            div_tags=soup.find_all('div',attrs={'class':'et_pb_team_member_description'})
            for tag in div_tags:
                job_txt = tag.find('p',attrs={'class':'et_pb_member_position'}).text
                if 'Support Services' in job_txt or 'Counselor' in job_txt:
                    name = tag.find('h4',attrs={'class':'et_pb_module_header'}).text
                    a_tag=tag.find('a')
                    if a_tag:
                        contact_info[school][name]=a_tag.get('href')
                        continue
                    p_tag=tag.find('p',attrs={'class':None})
                    if p_tag:
                        txt=p_tag.text
                        at_symb_ind = txt.find('@')
                        if at_symb_ind!=-1:
                            i = at_symb_ind-1
                            while txt[i-1]!=' ' and txt[i-1]!='\n':
                                i-=1
                            contact_info[school][name]=txt[i:txt.find('.org')+4]
        elif school=='University Preparatory School':
            soup=get_soup('https://www.uprepmilliones.com/staff')
            div_tags=soup.find_all('div',attrs={'role':'listitem'})
            for tag in div_tags:
                job_tag=tag.find('span',attrs={'style':'font-family:futura-lt-w01-book,sans-serif;'})
                if job_tag:
                    job_txt=job_tag.text
                    if ('PSE' in job_txt or 'Counselor' in job_txt or 'Support' in job_txt or 'Psychologist' in job_txt) and 'MS' not in job_txt:
                        name = ' '.join([tag.text for tag in tag.find_all('span',attrs={'style':'font-family:open sans,sans-serif;'})])
                        contact_info[school][name.replace('\xa0','')] = tag.find('a').text
        elif school=='Urban Pathways Charter School':
            soup=get_soup(school_to_link[school],'counseling-and-support-services/')
            p_tags=soup.find_all('p',attrs={'style':re.compile('padding')})
            for tag in p_tags:
                tag_txt = tag.text
                if ('Counselor' in tag_txt and 'Internship' not in tag_txt) or 'Psychologist' in tag_txt:
                    a_tag=tag.find('a')
                    contact_info[school][a_tag.text]=a_tag.get('href')
        elif school=='Woodland Hills High School':
            soup=get_soup('https://www.whsd.net/departments/special-education')
            header_tags=soup.find('div',attrs={'class':'fsElement fsContent'}).find('div',attrs={'class':'fsElementContent'}).find_all(re.compile('h[5-6]'))
            for i in range(0,len(header_tags),3):
                strong_tag = header_tags[i].find('strong')
                if strong_tag:
                    name_and_pos_txt = strong_tag.text
                    if 'Special Education' in name_and_pos_txt:
                        email_txt = header_tags[i+1].text
                        contact_info[school][name_and_pos_txt[:name_and_pos_txt.find('\n')]]=email_txt[email_txt.find('Email: ')+7:]
    return contact_info

def get_adams_contacts():
    """
    Retrieves the contact info for high school counselors and special education specialists in Philadelphia County
    """
    contact_info = dict()
    school_to_link=get_school_to_link('adams_hs_links.json','Adams')
    for school in school_to_link:
        contact_info[school]=dict()
        if school=='Biglerville High School':
            soup=get_soup('https://www.upperadams.org/departments/student-services-special-education/department-home/special-education')
            tags = soup.find_all(re.compile('em|p'))
            num_tags = len(tags)
            i=0
            while i<num_tags and tags[i].text!='Biglerville High School':
                i+=1
            i+=1
            while i<num_tags and tags[i].name!='em':
                tag_txt = tags[i].text
                if 'Learning' in tag_txt or 'Autistic' in tag_txt:
                    contact_info[school][tag_txt[:tag_txt.rfind('-')-1]]=tags[i].find('a').text.strip('mailto:')
                i+=1
        elif school=='Littlestown High School':
            soup=get_soup(school_to_link[school],'/staff?filter_ids=117586,117596')
            div_tags=soup.find_all('div',attrs={'class':'staff-info'})
            for tag in div_tags:
                dept = tag.find('div',attrs={'class':'department'}).text
                if 'Learning Support' in dept or 'Guidance' in dept:
                    name = tag.find('div',attrs={'class':'name'}).text
                    comma_ind = name.find(',')
                    if comma_ind==-1:
                        comma_ind=len(name)-1
                    contact_info[school][name[:comma_ind].strip('\n ')]=tag.find('a').text.strip('\n ')
    return contact_info

def get_armstrong_contacts():
    """
    Retrieves the contact info for high school counselors and special education specialists in Armstrong County
    """
    contact_info = dict()
    school_to_link=get_school_to_link('armstrong_hs_links.json','Armstrong')
    for school in school_to_link:
        contact_info[school]=dict()
        if school=='Apollo-Ridge High School':
            soup=get_soup(school_to_link[school],'apps/departments/')
            spec_ed_dept_url_suff = soup.find('a',string='Special Education').get('href')[1:]
            spec_ed_dept_url_suff = spec_ed_dept_url_suff+'&pREC_ID=staff'
            soup=get_soup(school_to_link[school],spec_ed_dept_url_suff)
            div_tags=soup.find_all('div',attrs={'class':'user-info ada'})
            for tag in div_tags:
                contact_info[school][tag.find('a',attrs={'class':'name'}).text.strip('\n\t')]=tag.find('span',attrs={'class':'user-email'}).find('a',attrs={'class':'email'}).text
    return contact_info

def get_bedford_contacts():
    contact_info=dict()
    school_to_link=get_school_to_link('bedford_hs_links.json','Bedford')
    for school in school_to_link:
        contact_info[school]=dict()
        if school=='Bedford High School':
            get_contacts_from_staff_info_blocks(school_to_link,school,['Learning Support','Special Ed','Counselor'],contact_info)
        elif school=='Chestnut Ridge Senior High School':
            soup=get_soup(school_to_link[school],'apps/staff')
            div_tags=soup.find_all('div',attrs={'class':'user-info ada'})
            for tag in div_tags:
                job=tag.find('span',attrs={'class':'user-position user-data'})
                if job and 'Special Ed' in job.text:
                    contact_info[school][tag.find('a',attrs={'class':'name'}).text.strip('\n\t')]=tag.find('span',attrs={'class':'user-email'}).find('a',attrs={'class':'email'}).text
        elif school=='HOPE for Hyndman Charter School':
            soup=get_soup(school_to_link[school],'staff-1')
            p_tags=soup.find_all('p')
            for tag in p_tags:
                tag_txt=tag.text
                if 'Special Ed' in tag_txt:
                    contact_info[school][tag_txt[:tag_txt.find('\xa0')]]=tag.find('a').text
        elif school=='Tussey Mountain Junior/Senior High School':
            get_contacts_from_staff_info_blocks(school_to_link,school,['Support'],contact_info)  
    return contact_info

def get_blair_contacts():
    contact_info=dict()
    school_to_link=get_school_to_link('blair_hs_links.json','Blair')
    for school in school_to_link:
        contact_info[school]=dict()
        if school=='Williamsburg Community Junior/Senior High School':
            soup=get_soup(school_to_link[school])
            soup=get_soup(school_to_link[school]+soup.find('span',string='Special Education').parent.get('href')[1:])
            div_tags=soup.find_all('div',attrs={'class':None,'id':None})
            i = 0
            n = len(div_tags)
            while i<n:
                name_tag = div_tags[i].find('strong')
                if name_tag:
                    job_txt = div_tags[i+1].text
                    if 'Special Ed' in job_txt or 'Psychologist' in job_txt or 'Learning Support 9-12' in job_txt:
                        comma_ind = name_tag.text.find(',')
                        if comma_ind==-1:
                            comma_ind = len(name_tag.text)
                        contact_info[school][name_tag.text[:comma_ind]]=div_tags[i+2].text.strip('\xa0')
                    i+=2
                i+=1
        elif school=='Bishop Guilfoyle High School':
            soup=get_soup(school_to_link[school],'index.php?page=guidance-counseling')
            contact_col=soup.find('div',attrs={'class':'twoColumnLeft'})
            div_tags=contact_col.find_all('div')
            for tag in div_tags:
                tag_txts = tag.text.split(')')
                for txt in tag_txts:
                    if 'Counselor' in txt and is_hs_staff(txt):
                        end_name_ind = txt.find(' ', txt.find(' ')+1)
                        end_email_ind = txt.find(' ',end_name_ind+1)
                        contact_info[school][txt[:end_name_ind].strip(',')]=txt[end_name_ind+1:end_email_ind].strip(',')
        elif school=='Greater Altoona Career and Technology Center':
            soup=get_soup(school_to_link[school],'student-services/our-staff/')
            div_tags=soup.find_all('div',attrs={'class':'abcfslPadLPc5 abcfslCenter575'})
            for tag in div_tags:
                job_txt = tag.find('div',attrs={'class':'dirTitle'}).text
                if 'Counselor' in job_txt or 'Special Population' in job_txt:
                    contact_info[school][tag.find('h3').text]=tag.find('a').text
        elif school=='Tyrone Area High School':
            soup=get_soup(school_to_link[school])
            soup=get_soup(school_to_link[school]+soup.find('span',string='Departments').parent.get('href')[1:])
            soup=get_soup(school_to_link[school]+soup.find('a',string='Student Services').get('href')[1:])
            soup=get_soup(school_to_link[school]+soup.find('span',string='Special Education').parent.get('href')[1:])
            a_tags = soup.find_all('a',attrs={'id':None,'title':None,'href':re.compile('([A-Za-z])*@([A-Za-z0-9])*.')})
            for tag in a_tags:
                contact_info[school][tag.text.strip('\xa0')]=tag.get('href').strip('mailto:')
    return contact_info

def write_to_excel_file(contact_info:dict,county:str,file_name:str):
    """
    Writes the content of contact_info to an Excel file 
    @param contact_info : dictionary mapping school name indicated by school to a dictionary mapping school staff member name to email
    @param county
    @param file_name
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = county
    sheet.cell(row=1, column=1).value='School'
    sheet.cell(row=1,column=2).value='Counselor Name'
    sheet.cell(row=1, column=3).value='Email'
    i = 2
    for school in contact_info:
        for couns in contact_info[school]:
            sheet.cell(row=i,column=1).value=school
            sheet.cell(row=i,column=2).value=couns
            sheet.cell(row=i,column=3).value=contact_info[school][couns]
            i+=1
    wb.save(file_name)

#get_state_hs_links('https://en.wikipedia.org/wiki/List_of_high_schools_in_Pennsylvania','Blair')
get_state_hs_links('https://en.wikipedia.org/wiki/List_of_high_schools_in_New_York','Allegany')
#print(get_blair_contacts())
#write_to_excel_file(get_blair_contacts(),'Blair','counselor_contacts.xlsx')
